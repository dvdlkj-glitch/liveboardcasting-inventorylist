"""Parse the raw customer-order PDFs and the stock Excel into two clean
import files (orders_import.xlsx, stock_import.xlsx)."""
import re
import pdfplumber
import openpyxl
from openpyxl import Workbook

UP = "/sessions/lucid-ecstatic-thompson/mnt/uploads"
OUT = "/sessions/lucid-ecstatic-thompson/mnt/outputs"

inv_re  = re.compile(r"Invoice\s*#\s*([A-Z0-9\-]+)")
date_re = re.compile(r"^\d{4}-\d{2}-\d{2}\s")
item_re = re.compile(r"^\s*(\d+)\.\s*\[([^\]]+)\]\s*-?\s*(.*)$")
qty_re  = re.compile(r"(\d+)\s*[xX]\s*([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$")
cust_re = re.compile(r"^(.*?)\s*\|\s*FB:")

def clean(s):
    return re.sub(r"\s+", " ", (s or "").strip())

def finalize(orders, inv, cust, p, q):
    desc = clean(" ".join(d for d in p["desc"] if d))
    orders.append({
        "invoice": inv, "customer": cust, "item_code": p["code"],
        "description": desc, "qty_ordered": int(q.group(1)),
        "unit_price": float(q.group(2).replace(",", "")),
        "line_total": float(q.group(3).replace(",", "")),
    })

def parse_orders():
    PDFS = ["slist 1.pdf", "slist2.pdf", "slist 3.pdf"]
    orders = []
    for fn in PDFS:
        with pdfplumber.open(f"{UP}/{fn}") as pdf:
            lines = []
            for pg in pdf.pages:
                lines += (pg.extract_text() or "").split("\n")
        cur_inv, cur_cust = "", ""
        pending = None
        for raw in lines:
            ln = raw.strip()
            m = inv_re.search(ln)
            if m:
                cur_inv = m.group(1); cur_cust = ""
            mc = cust_re.match(ln)
            if mc and not cur_cust:
                cur_cust = clean(mc.group(1))
            mi = item_re.match(ln)
            if mi:
                pending = {"code": mi.group(2).strip(), "desc": [clean(mi.group(3))]}
                q = qty_re.search(ln)
                if q:
                    finalize(orders, cur_inv, cur_cust, pending, q); pending = None
                continue
            if pending is not None:
                q = qty_re.search(ln)
                if q:
                    pre = qty_re.sub("", ln).strip()
                    if pre and not date_re.match(pre) and not re.match(r"^\d{1,2}/\d{1,2}/\d{2}", pre):
                        pending["desc"].append(clean(pre))
                    finalize(orders, cur_inv, cur_cust, pending, q); pending = None
                elif ln and "http" not in ln and not date_re.match(ln) and not re.match(r"^\d{1,2}/\d{1,2}/\d{2}", ln):
                    pending["desc"].append(clean(ln))
    return orders

def parse_stock():
    wb = openpyxl.load_workbook(f"{UP}/China Stock Listing.xlsx", data_only=True)
    rows = []
    def num(v):
        try: return int(float(v))
        except: return None
    # --- suk (SukGarden) header at row idx1 ---
    if "suk" in wb.sheetnames:
        ws = wb["suk"]
        for r in list(ws.iter_rows(values_only=True))[2:]:
            name = clean(str(r[3])) if r[3] else ""
            code = clean(str(r[6])) if r[6] else ""
            qty  = num(r[10])
            bar  = clean(str(r[7])) if r[7] else ""
            if name and qty:
                rows.append({"supplier":"SukGarden 蔬果园","item_code":code,
                             "description":name,"qty_received":qty,"barcode":bar})
    # --- 内衣 (图兰朵) header row0 ---
    if "内衣" in wb.sheetnames:
        ws = wb["内衣"]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            code = clean(str(r[1])) if r[1] else ""
            name = clean(str(r[2])) if r[2] else ""
            qty  = num(r[5])
            if name and qty:
                rows.append({"supplier":"图兰朵 Turandot","item_code":code,
                             "description":name,"qty_received":qty,"barcode":""})
    # --- 珠宝 (中国珠宝) name col1, qty col2 ---
    if "珠宝" in wb.sheetnames:
        ws = wb["珠宝"]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            name = clean(str(r[1])) if r[1] else ""
            qty  = num(r[2])
            if name and qty:
                rows.append({"supplier":"中国珠宝 China Jewelry","item_code":"",
                             "description":name,"qty_received":qty,"barcode":""})
    # --- 牛蛙 (xfrog) header row0: 序号,品牌,img,商品名称,规格,机制,采购 ---
    if "牛蛙" in wb.sheetnames:
        ws = wb["牛蛙"]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            code = clean(str(r[0])) if r[0] else ""
            name = clean(str(r[3])) if r[3] else ""
            qty  = num(r[6])
            if name and qty:
                rows.append({"supplier":"xfrog 华大牛蛙","item_code":code,
                             "description":name,"qty_received":qty,"barcode":""})
    return rows

def write_xlsx(path, headers, rows, keys):
    wb = Workbook(); ws = wb.active; ws.title = "import"
    ws.append(headers)
    for r in rows:
        ws.append([r.get(k, "") for k in keys])
    # autosize-ish
    for ci, h in enumerate(headers, 1):
        width = max(len(str(h)), *(len(str(r.get(keys[ci-1], ""))) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max(width+2, 10), 60)
    wb.save(path)

if __name__ == "__main__":
    orders = parse_orders()
    stock  = parse_stock()
    print("orders parsed:", len(orders))
    print("invoices:", sorted(set(o["invoice"] for o in orders)))
    print("customers:", sorted(set(o["customer"] for o in orders if o["customer"])))
    print("stock rows:", len(stock))
    from collections import Counter
    print("stock by supplier:", Counter(s["supplier"] for s in stock))
    print("\nsample orders:")
    for o in orders[:6]:
        print("  ", o)
    print("sample stock:")
    for s in stock[:6]:
        print("  ", s)

    for d in (OUT, "/sessions/lucid-ecstatic-thompson/mnt/Ulive purchase Order to Automation Process"):
        write_xlsx(f"{d}/orders_import.xlsx",
                   ["invoice","customer","item_code","description","qty_ordered","unit_price","line_total"],
                   orders,
                   ["invoice","customer","item_code","description","qty_ordered","unit_price","line_total"])
        write_xlsx(f"{d}/stock_import.xlsx",
                   ["supplier","item_code","description","qty_received","barcode"],
                   stock,
                   ["supplier","item_code","description","qty_received","barcode"])
    print("\nwrote orders_import.xlsx and stock_import.xlsx")
